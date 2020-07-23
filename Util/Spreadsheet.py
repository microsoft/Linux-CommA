# Copyright (c) Microsoft Corporation.
# Licensed under the MIT License.
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Set, Tuple

import git
import openpyxl
import sqlalchemy
from openpyxl.cell import cell
from openpyxl.workbook import workbook
from openpyxl.worksheet import worksheet

import Util.Config
import Util.Tracking
from DatabaseDriver.DatabaseDriver import DatabaseDriver
from DatabaseDriver.SqlClasses import Distros, MonitoringSubjects, PatchData
from UpstreamTracker.ParseData import process_commits


def get_db_commits() -> Dict[str, int]:
    """Query the 'PatchData' table for all commit hashes and IDs."""
    with DatabaseDriver.get_session() as s:  # type: sqlalchemy.orm.session.Session
        return {
            commit_id: patch_id
            for (commit_id, patch_id) in s.query(
                PatchData.commitID, PatchData.patchID
            ).filter(
                # Exclude ~1000 CIFS patches.
                ~PatchData.affectedFilenames.like("%fs/cifs%")
            )
        }


def get_workbook(in_file: str) -> Tuple[workbook.Workbook, worksheet.Worksheet]:
    """Open the spreadsheet and return it and the 'git log' worksheet.

    Also fix the pivot table so the spreadsheet doesn't crash.

    """
    if not Path(in_file).exists():
        logging.error(f"The file {in_file} does not exist")
        sys.exit(1)
    wb = openpyxl.load_workbook(filename=in_file)
    # Force refresh of pivot table in “Pivot” worksheet.
    pivot = wb["Pivot"]._pivots[0]
    pivot.cache.refreshOnLoad = True
    # The worksheet is manually named “git log”.
    ws = wb["git log"]
    return (wb, ws)


def get_column(ws: worksheet.Worksheet, name: str) -> cell.Cell:
    """Gets the header cell for the given column name.

    The column names are the values of the cells in the first row,
    e.g. 'Commit Title'. Use 'cell.column' for the numeric column, or
    'cell.column_letter' for the letter. Given a row, index into it
    with 'row[get_column(ws, "Commit Title").column]'.

    """
    return next(c for c in ws[1] if c.value == name)


def get_wb_commits(ws: worksheet.Worksheet) -> Set[str]:
    """Get every commit in the workbook."""
    # Skip the header and all ‘None’ values.
    column = get_column(ws, "Commit ID").column_letter
    return {c.value for c in ws[column][1:] if c.value is not None}


def import_commits(in_file: str) -> None:
    """This adds commits from the spreadsheet into the database.

    This lets us track additional commits that were manually added to
    the spreadsheet, but were not automatically found by CommA's
    upstream monitoring logic.

    """
    print(f"Importing commits from spreadsheet '{in_file}'...")
    wb, ws = get_workbook(in_file)
    wb_commits = get_wb_commits(ws)
    db_commits = get_db_commits()
    missing_commits = wb_commits - db_commits.keys()
    process_commits(commit_ids=missing_commits, add_to_database=True)
    print("Finished importing!")


def include_commit(sha: str, repo: git.Repo, base_commit: git.Commit) -> bool:
    """Determine if we should export the commit."""
    # Skip empty values (such as if ‘cell.value’ was passed).
    if sha is None:
        return False
    # Skip commits that aren’t in the repo.
    try:
        commit = repo.commit(sha)
    except ValueError:
        return False
    # Skip commits before the chosen base.
    if not repo.is_ancestor(base_commit, commit):
        return False
    # Skip commits to tools.
    filenames = Util.Tracking.get_filenames(commit)
    if any(f.startswith("tools/hv/") for f in filenames):
        return False
    return True


def create_commit_row(
    sha: str, repo: git.Repo, ws: worksheet.Worksheet
) -> Dict[str, Any]:
    """Create a row with the commit's SHA, date, release and title."""
    commit = repo.commit(sha)
    # TODO: Some (but not all) of this info is available in the
    # database, so if add the release to the database we can skip
    # using the commit here.
    date = datetime.utcfromtimestamp(commit.authored_date).date()
    title = commit.message.split("\n")[0]
    # Get the ‘v5.7’ from ‘v5.7-rc1-2-gc81992e7f’.
    # NOTE: This must be ordered “--contains <SHA>” for Git.
    tag = repo.git.describe("--contains", sha)
    release = re.search(r"(v[^-~]*)[-~]", tag).group(1)

    # The worksheet has additional columns with manually entered
    # info, which we can’t insert, so we skip them.
    def get_letter(name: str) -> str:
        return get_column(ws, name).column_letter

    return {
        get_letter("Commit ID"): sha,
        get_letter("Date"): date,
        get_letter("Release"): release,
        get_letter("Commit Title"): title[: min(len(title), 120)],
    }


def export_commits(in_file: str, out_file: str) -> None:
    """This adds commits from the database to the spreadsheet.

    This lets us automatically update the spreadsheet by adding
    commits which CommA found. It adds the basic information available
    from the commit.

    """
    wb, ws = get_workbook(in_file)
    wb_commits = get_wb_commits(ws)

    # Collect the commits in the database which are not in the
    # workbook, but that we want to include.
    db_commits = get_db_commits()
    repo = Util.Tracking.get_linux_repo()
    base_commit = repo.commit("v4.11")
    missing_commits = [
        commit
        for commit in list(db_commits.keys() - wb_commits)
        if include_commit(commit, repo, base_commit)
    ]

    # Append each missing commit as a new row to the commits
    # worksheet.
    print(f"Exporting {len(missing_commits)} commits to {out_file}...")
    for commit in missing_commits:
        # TODO: Set fonts of the cells.
        ws.append(create_commit_row(commit, repo, ws))

    wb.save(out_file)
    print("Finished exporting!")


def get_distros() -> List[str]:
    """Collect the distros we’re tracking in the database."""
    with DatabaseDriver.get_session() as s:
        # TODO: Handle Debian.
        return [d for (d,) in s.query(Distros.distroID) if not d.startswith("Debian")]


def get_fixed_patches(commit: str, commits: Dict[str, int]) -> str:
    """Get the fixed patches for the given commit."""
    with DatabaseDriver.get_session() as s:
        patch = s.query(PatchData).filter_by(patchID=commits[commit]).one()
        return patch.fixedPatches


def get_revision(distro: str, commit: str, commits: Dict[str, int]) -> str:
    """Get the kernel revision which includes commit or 'Absent'."""
    # NOTE: For some distros (e.g. Ubuntu), we continually add new
    # revisions (Git tags) as they become available, so we need the
    # max ID, which is the most recent.
    with DatabaseDriver.get_session() as s:
        subject, _ = (
            s.query(
                MonitoringSubjects,
                sqlalchemy.func.max(MonitoringSubjects.monitoringSubjectID),
            )
            .filter_by(distroID=distro)
            .one()
        )

        # TODO: We could try to simplify this using the
        # ‘monitoringSubject’ relationship on the ‘PatchData’ table,
        # but because the database tracks what’s missing, it becomes
        # hard to state where the patch is present.
        missing_patch = subject.missingPatches.filter_by(
            patchID=commits[commit]
        ).one_or_none()

        if missing_patch is None:  # Then it’s present.
            return subject.revision
        else:
            return "Absent"


def update_commits(in_file: str, out_file: str) -> None:
    """Update each row with the 'Fixes' and distro information."""
    wb, ws = get_workbook(in_file)
    commits = get_db_commits()
    distros = get_distros()
    commit_column = get_column(ws, "Commit ID").column_letter

    for commit_cell in ws[commit_column][1:]:  # Skip the header row
        commit = commit_cell.value
        if commit is None:
            continue  # Ignore empty rows.

        def get_cell(name: str) -> cell.Cell:
            """Get the cell of the named column in this commit's row."""
            return ws[f"{get_column(ws, name).column_letter}{commit_cell.row}"]

        # Update “Fixes” column.
        if commit in commits:
            get_cell("Fixes").value = get_fixed_patches(commit, commits)

        # Update all distro columns.
        #
        # TODO: Check that each distro is in the header row.
        for distro in distros:
            if commit in commits:
                get_cell(distro).value = get_revision(distro, commit, commits)
            else:
                get_cell(distro).value = "Unknown"

    wb.save(out_file)
    print("Finished updating!")
