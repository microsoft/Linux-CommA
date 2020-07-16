# Copyright (c) Microsoft Corporation.
# Licensed under the MIT License.
import logging
import re
import sqlalchemy
from datetime import datetime
from pathlib import Path

import git
import openpyxl
import Util.Config
import Util.Tracking

from DatabaseDriver.DatabaseDriver import DatabaseDriver
from DatabaseDriver.SqlClasses import (
    PatchData,
    Distros,
    MonitoringSubjects,
    MonitoringSubjectsMissingPatches,
)

Util.Config.dry_run = True

# Work in progress to export this data to a particular spreadsheet.
# Currently just comparing an existing spreadsheet and our data.
def export_spreadsheet(in_file: str, out_file: str):
    if not Path(in_file).exists():
        logging.error(f"The file {in_file} does not exist")
        return

    # There’s an empty row after the header, so skip the header and
    # all ‘None’ values. The sheet is manually named “git log”.
    wb = openpyxl.load_workbook(filename=in_file)
    ws = wb["git log"]
    wb_commits = {c.value for c in ws["A"][1:] if c.value is not None}

    db_commits = set()
    with DatabaseDriver.get_session() as s:  # type: sqlalchemy.orm.session.Session
        db_commits = {
            c
            for c, in s.query(PatchData.commitID).filter(
                # Exclude ~1000 CIFS patches.
                ~PatchData.affectedFilenames.like("%fs/cifs%")
            )
        }

    repo = Util.Tracking.get_repo()  # type: git.Repo
    base_commit = repo.commit("v4.11")

    def include_commit(sha):
        try:
            commit = repo.commit(sha)
            # Skip commits before the chosen base.
            if not repo.is_ancestor(base_commit, commit):
                return False
            # Skip commits to tools.
            filenames = Util.Tracking.get_filenames(commit)
            if any(f.startswith("tools/hv/") for f in filenames):
                return False
            return True
        except ValueError:
            return False

    # Filter to desired missing commits which we want to append.
    commits = [sha for sha in list(db_commits - wb_commits) if include_commit(sha)]

    def commit_to_row(sha):
        commit = repo.commit(sha)
        date = datetime.utcfromtimestamp(commit.authored_date).strftime("%Y-%m-%d")
        summary = commit.message.split("\n")[0]
        # Get the ‘v5.7’ from ‘v5.7-rc1-2-gc81992e7f’
        tag = repo.git.describe(sha, contains=None)
        release = re.search(r"(v[^-~]*)[-~]", tag).group(1)
        return {
            "A": sha,
            "B": date,
            "C": release,
            "G": summary[: min(len(summary), 120)],
        }

    for commit in commits:
        print(f"Appending {commit}...")
        ws.append(commit_to_row(commit))
    print(f"Exported {len(commits)} commits to {out_file}!")

    ws = wb.create_sheet("distros")
    with DatabaseDriver.get_session() as s:
        distros = [d for (d,) in s.query(Distros.distroID)]

    ws.append(["Commit ID"] + distros)

    def commit_to_status(commit):
        row = [commit]
        with DatabaseDriver.get_session() as s:
            for distro in distros:
                # TODO: This logic awful because the database schema
                # is a bit obtuse. We need to find the max subject ID
                # for the distro (meaning the most recently added
                # revision, so the latest code), then find the patch
                # ID for the given commit, then check if that patch ID
                # is in the list of missing patches for the subject
                # ID, as stored in `MonitoringSubjectsMissingPatches`.

                (subject_id,) = (
                    s.query(sqlalchemy.func.max(MonitoringSubjects.monitoringSubjectID))
                    .filter_by(distroID=distro)
                    .one()
                )
                (patch_id,) = (
                    s.query(PatchData.patchID).filter_by(commitID=commit).one()
                )
                listed_missing = (
                    s.query(MonitoringSubjectsMissingPatches)
                    .filter_by(monitoringSubjectID=subject_id)
                    .filter_by(patchID=patch_id)
                    .one_or_none()
                )
                if listed_missing is None:  # Then it’s present.
                    row.append("Present")
                else:
                    row.append("Absent")
            return row

    for commit in db_commits:
        ws.append(commit_to_status(commit))

    # Save spreadsheet.
    wb.save(out_file)
