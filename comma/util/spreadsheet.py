# Copyright (c) Microsoft Corporation.
# Licensed under the MIT License.
"""
Functions for exporting data to Excel spreadsheets
"""

import logging
import re
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import git
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.formula.translate import Translator
from openpyxl.styles import DEFAULT_FONT
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from comma.database.model import MonitoringSubjects, MonitoringSubjectsMissingPatches, PatchData
from comma.exceptions import CommaSpreadsheetError


LOGGER = logging.getLogger(__name__)


class WorksheetWrapper:
    """
    Wrapper for openpyxl worksheet
    Consolidates common worksheet logic
    """

    def __init__(self, worksheet: Worksheet) -> None:
        self.worksheet = worksheet

    @lru_cache  # noqa: B019
    def get_column(self, name: str) -> int:
        """Get column by letter"""
        return next(cell for cell in self.worksheet[1] if cell.value == name).column

    def get_column_cells(self, name: str) -> List[Cell]:
        """Get cells for a specified column if the cell has a value"""

        column = self.get_column(name)
        return tuple(
            cell
            for cell in next(self.worksheet.iter_cols(min_row=2, min_col=column, max_col=column))
            if cell.value is not None
        )

    def get_cell(self, name: str, row) -> Cell:
        """Get the cell of the named column in this commit's row."""
        return self.worksheet.cell(row, self.get_column(name))

    def append(self, row: Dict[str, Any]) -> None:
        """
        Given a dictionary with column headers as rows, format and append to worksheet
        """
        self.worksheet.append({self.get_column(key): value for key, value in row.items()})

    def sort(self, column=2, key=None, reverse=True):
        """
        Sort spreadsheet based on specified column
        Drops row that do not have value for column
        Assumes the first row is a header
        """
        idx = column - 1

        if key is None:

            def key(row):
                value = row[idx].value
                if isinstance(value, datetime):
                    value = value.date()
                return value

        # Sort rows
        rows = sorted(
            (row for row in self.worksheet.iter_rows(min_row=2) if row[idx].value is not None),
            key=key,
            reverse=reverse,
        )

        # Clear worksheet
        self.worksheet.delete_rows(2, self.worksheet.max_row - 1)

        # Clean up values
        for row_num, row in enumerate(rows, 2):
            for old_cell in row:
                new_cell = self.worksheet.cell(row_num, old_cell.column, old_cell.value)
                new_cell.font = DEFAULT_FONT

                # Fix dates so they get stored properly
                if isinstance(old_cell.value, datetime) and not any(
                    (old_cell.value.hour, old_cell.value.minute, old_cell.value.second)
                ):
                    new_cell.value = old_cell.value.date()
                    new_cell.number_format = "YYYY-MM-DD"

                # If value is a formula, translate it
                if old_cell.data_type == "f":
                    new_cell.value = Translator(
                        new_cell.value, old_cell.coordinate
                    ).translate_formula(new_cell.coordinate)


def get_workbook(in_file: str) -> Tuple[Workbook, WorksheetWrapper]:
    """Open the spreadsheet and return it and the 'git log' worksheet.

    Also fix the pivot table so the spreadsheet doesn't crash.

    """
    if not Path(in_file).exists():
        raise CommaSpreadsheetError(f"Input spreadsheet '{in_file}' does not exist")

    workbook = openpyxl.load_workbook(filename=in_file)

    # Force refresh of pivot table in “Pivot” worksheet.
    LOGGER.debug("Finding worksheet named 'Pivot'...")
    pivot = workbook["Pivot"]._pivots[0]  # pylint: disable=protected-access
    pivot.cache.refreshOnLoad = True

    # The worksheet is manually named “git log”.
    LOGGER.debug("Finding worksheet named 'git log'...")
    worksheet = workbook["git log"]

    return (workbook, WorksheetWrapper(worksheet))


class Spreadsheet:
    """
    Parent object for symbol operations
    """

    def __init__(self, config, database, repo) -> None:
        self.config = config
        self.database = database
        self.repo = repo

    def get_db_commits(
        self, since: Optional[float] = None, excluded_paths: Optional[Iterable[str]] = None
    ) -> Dict[str, int]:
        """Query the 'PatchData' table for all commit hashes and IDs."""
        with self.database.get_session() as session:
            query = session.query(PatchData.commitID, PatchData.patchID)

            if excluded_paths:
                for entry in excluded_paths:
                    query = query.filter(~PatchData.affectedFilenames.like(entry))

            if since:
                query = query.filter(PatchData.commitTime >= since)

            return dict(query)

    def get_release(self, sha: str) -> str:
        """Get the ‘v5.7’ from ‘v5.7-rc1-2-gc81992e7f’."""
        try:
            # NOTE: This must be ordered “--contains <SHA>” for Git.
            tag = self.repo.git.describe("--contains", sha)
            # Use "(v[^-~]+(-rc[0-9]+)?)[-~]" to include ‘-rcX’  # pylint: disable=wrong-spelling-in-comment
            return re.search(r"(v[^-~]*)[-~]", tag)[1]
        except git.GitCommandError:
            return "N/A"

    def export_commits(self, in_file: str, out_file: str) -> None:
        """This adds commits from the database to the spreadsheet.

        This lets us automatically update the spreadsheet by adding commits which CommA found.
        It adds the basic information available from the commit.

        """
        workbook, worksheet = get_workbook(in_file)

        # Get commits in database, but not in spreadsheet
        # Exclude ~1000 CIFS patches and anything that touches tools/hv  # pylint: disable=wrong-spelling-in-comment
        missing_commits = self.get_db_commits(
            since=self.config.upstream_since.datetime,
            excluded_paths=self.config.spreadsheet.excluded_paths,
        ).keys() - {cell.value for cell in worksheet.get_column_cells("Commit ID")}

        exported = 0
        to_export = len(missing_commits)
        LOGGER.info("Exporting %d commits to %s", to_export, out_file)

        # Append each missing commit as a new row to the commits worksheet.
        for commit_id in missing_commits:
            if commit_id is None:
                LOGGER.error("Commit in database has an empty commit ID")
                continue

            # Skip commits that are not in the repo.
            try:
                commit = self.repo.commit(commit_id)
            except ValueError:
                LOGGER.warning("Commit '%s' not in repo!", commit_id)
                continue

            # TODO (Issue 40): If release was added to the database, commit could be skipped and
            # all data could be pulled from the database
            worksheet.append(
                {
                    "Commit ID": commit_id,
                    "Date": datetime.utcfromtimestamp(commit.authored_date).date(),
                    "Release": self.get_release(commit_id),
                    "Commit Title": "{:.120}".format(commit.message.split("\n")[0]),
                }
            )

            # Periodically report status in case we have a lot of commits
            exported += 1
            if exported and not exported % 50:
                LOGGER.info("Exported %d of %d commits", exported, to_export)

        LOGGER.info("%d commits exported to %s", exported, out_file)
        worksheet.sort()
        workbook.save(out_file)
        LOGGER.info("Finished exporting!")

    def update_commits(self, in_file: str, out_file: str) -> None:
        # pylint: disable=too-many-locals
        """Update each row with the 'Fixes' and distro information."""
        workbook, worksheet = get_workbook(in_file)
        # Exclude ~1000 CIFS patches
        db_commits = self.get_db_commits(excluded_paths=self.config.spreadsheet.excluded_paths)
        targets = {}

        with self.database.get_session() as session:
            for repo in self.database.get_downstream_repos():
                # TODO (Issue 51): Handle Debian
                if repo.startswith("Debian"):
                    continue

                # Make sure there is a column in the spreadsheet
                try:
                    worksheet.get_column(repo)
                except StopIteration as e:
                    raise CommaSpreadsheetError(
                        f"No column with distro '{repo}', please fix spreadsheet!"
                    ) from e

                # Get the latest monitoring subject for the remote
                targets[repo] = (
                    session.query(
                        MonitoringSubjects.monitoringSubjectID, MonitoringSubjects.revision
                    )
                    .filter_by(distroID=repo)
                    .order_by(MonitoringSubjects.monitoringSubjectID.desc())
                    .limit(1)
                    .one()
                )

            commits_cells = worksheet.get_column_cells("Commit ID")
            total_rows = len(commits_cells)
            LOGGER.info("Evaluating updates for %d rows", total_rows)

            # Iterate through commit IDs in spreadsheet. Skip the header row.
            for count, commit_cell in enumerate(commits_cells):
                if count and not count % 50:
                    LOGGER.info("Evaluated updates for %d of %d rows", count, total_rows)

                patch_id = db_commits.get(commit_cell.value)

                # If patch isn't in the database, set all distros to unknown
                if patch_id is None:
                    for distro in targets:
                        worksheet.get_cell(distro, commit_cell.row).value = "Unknown"
                    continue

                # Update “Fixes” column.
                patch = session.query(PatchData).filter_by(patchID=patch_id).one()
                # The database stores these separated by a space, but we want commas
                worksheet.get_cell("Fixes", commit_cell.row).value = (
                    ", ".join(patch.fixedPatches.split()) if patch.fixedPatches else None
                )

                # Update all distro columns.
                for distro, subject in targets.items():
                    # TODO (Issue 40): We could try to simplify this using the monitoringSubject
                    # relationship on the PatchData table, but because the database tracks
                    # what’s missing, it becomes hard to state where the patch is present.
                    missing_patch = (
                        session.query(MonitoringSubjectsMissingPatches.monitoringSubjectID)
                        .filter_by(patchID=patch_id)
                        .filter_by(monitoringSubjectID=subject.monitoringSubjectID)
                        .scalar()
                    )
                    worksheet.get_cell(distro, commit_cell.row).value = (
                        subject.revision if missing_patch is None else "Absent"
                    )

            LOGGER.info("Updates evaluated for %s rows", total_rows)

            workbook.save(out_file)
            LOGGER.info("Finished updating!")
